import openpyxl, json


def retrieve_parameters_cell_position_to_list(parameter_value, template_sheet) -> list:
        global sectors_parameters_position, project_parameters_position
        for row in template_sheet.iter_rows():
            for cell in row:
                if cell.value == parameter_value and 'sectors' in parameter_value: sectors_parameters_position = cell.coordinate
                elif cell.value == parameter_value and 'project' in parameter_value: project_parameters_position = cell.coordinate

        cell_position = sectors_parameters_position if 'sectors' in parameter_value else project_parameters_position
        cells_list = []
        for _ in range(100):
            cell_position = f'{cell_position[0]}{int(cell_position[1:])+1}'
            if template_sheet[cell_position].value is not None:
                cells_list.append(template_sheet[cell_position].value)
            else: break

        return cells_list


def record_composition_data(template_workbook, new_workbook, comp_id, comp_name, comp_unit, comp_proj_date, sector_dict):

    # switch to composition sheet in template file
    template_sheet = template_workbook['Composições']
    # create same sheet in the new workbook
    comp_sheet = new_workbook.create_sheet(title=template_sheet.title)
    # fill in composition details in new sheet
    comp_data = {'comp_id': comp_id, 'comp_name': comp_name, 'comp_unit': comp_unit, 'comp_proj_date': comp_proj_date}
    comp_cells_data = {}
    all_total = 0
    position_counter = 0

    for row in template_sheet.iter_rows():
        for cell in row:
            # insert normal parameters that are needed, such as heading
            if cell.value != None:
                comp_sheet[cell.coordinate] = cell.value
                if 'comp' not in cell.value.lower(): comp_cells_data.update({cell.value: cell.coordinate})
            if cell.value in comp_data:
                try: comp_sheet[cell.coordinate] = comp_data[cell.value]
                except ValueError: pass

    # fill in composition items data in sheet as per layout
    for sector in sector_dict:
        for comp_item in sector_dict[sector]:
            for cell_data in comp_cells_data:
                cell_position = f'{comp_cells_data[cell_data][0]}{int(comp_cells_data[cell_data][1:]) + position_counter}'
                if cell_data in comp_item: comp_sheet[cell_position] = f'P.{comp_item[cell_data]}' if cell_data == 'ID' else  comp_item[cell_data]
                if cell_data == 'bom_quant_avgprice':
                    bom_quant_avgprice = float(comp_item['bom_quant']) * float(comp_item['bom_avgprice'])
                    comp_sheet[cell_position] = bom_quant_avgprice
                    all_total += bom_quant_avgprice
                
                if len(sector_dict[sector]) - 1 == position_counter:
                    comp_sheet[f"{comp_cells_data['bom_quant'][0]}{int(comp_cells_data['bom_quant'][1:]) + position_counter + 1}"] = 'TOTAL MATERIAL/EQUIPAMENTO'
                    comp_sheet[f"{comp_cells_data['bom_quant_avgprice'][0]}{int(comp_cells_data['bom_quant_avgprice'][1:]) + position_counter + 1}"] = all_total
                
            # increment position to go to next row
            position_counter += 1


def record_item_data(template_workbook, new_workbook, prod_id, prod_name, prod_unit, items_dict, last_prod_position_value):

    # switch to item sheet in template file
    template_sheet = template_workbook['Cotações']
    # create same sheet in the new workbook
    prod_sheet = new_workbook.create_sheet(title=template_sheet.title) if last_prod_position_value == 0 else new_workbook[template_sheet.title]
    # fill in item details in new sheet
    prod_data = {'prod_id': prod_id, 'prod_name': prod_name, 'prod_unit': f'UND : {prod_unit}'}
    item_data = {} # {'EMPRESA': 'A7', 'CONTATO': 'B7', 'VALOR PROPOSTA (R$)': 'C7'}

    for row in template_sheet.iter_rows():
        for cell in row:
            # insert normal parameters that are needed, such as headings
            if cell.value != None:
                if last_prod_position_value == 0: prod_sheet[cell.coordinate] = cell.value

                if cell.value in ('EMPRESA', 'CONTATO', 'VALOR PROPOSTA (R$)'):
                    item_data.update({cell.value: cell.coordinate})
                    prod_sheet[f'{cell.coordinate[0]}{int(cell.coordinate[1:]) + last_prod_position_value}'] = cell.value

            if cell.value in prod_data:
                try: prod_sheet[f'{cell.coordinate[0]}{int(cell.coordinate[1:]) + last_prod_position_value}'] = prod_data[cell.value]
                except ValueError: pass
            
    last_prod_position_value += 1

    for item in items_dict:
        prod_sheet[f"{item_data['EMPRESA'][0]}{int(item_data['EMPRESA'][1:]) + last_prod_position_value}"] = f"{item['supplier_ID']} - {item['supplier']}"
        prod_sheet[f"{item_data['CONTATO'][0]}{int(item_data['CONTATO'][1:]) + last_prod_position_value}"] = f"{item['supplier_email']} - {item['supplier_address']} - {item['supplier_phone_1']}"
        if len(item['prices_history']) != 0:
            prices_history = sum([price[0] for price in item['prices_history']])
            prod_sheet[f"{item_data['VALOR PROPOSTA (R$)'][0]}{int(item_data['VALOR PROPOSTA (R$)'][1:]) + last_prod_position_value}"] = prices_history
        else: prod_sheet[f"{item_data['VALOR PROPOSTA (R$)'][0]}{int(item_data['VALOR PROPOSTA (R$)'][1:]) + last_prod_position_value}"] = None

        last_prod_position_value += 1

    print(prod_data)
    print(item_data)
    
    last_prod_position_value += 2

    return last_prod_position_value


last_prod_position_value = 0

def export_report(template_xls, proj_ID):
    global sectors_parameters_position, project_parameters_position, last_prod_position_value
    """receives the template.xls, the project_ID and the export_report file name to create an xls report.
    Saves the report.xls at S3 and return the file url."""

    # load project by proj_ID
    filename=f"jsons/{proj_ID}.json"
    with open(file=filename, mode='r') as file:
        projects_data = json.load(fp=file)['projects']

    # read from the templates file the parameters
    template_workbook = openpyxl.load_workbook(filename=f'templates/{template_xls}')

    template_sheet = template_workbook['Orçamento']
    # create a new worksheet
    new_workbook = openpyxl.Workbook()

    new_sheet = new_workbook.active
    # rename new_sheet
    new_sheet.title = template_sheet.title

    # take note of last cell position, for use later
    sector_position = None
    # use contents in template sheet to create contents in new sheet
    for row in template_sheet.iter_rows():
        for cell in row:
            # insert normal parameters that are needed, such as heading
            if cell.value != None:
                new_sheet[cell.coordinate] = cell.value
            # insert values of parameters that are present in json
            if cell.value in projects_data[0]:
                try: new_sheet[cell.coordinate] = projects_data[0][cell.value]
                except ValueError: pass
            # when finally at sectors, start restructuring data sector by sector, based on how many sectors are available
            if cell.value == 'sectors': sector_position = cell.coordinate
            
    if sector_position is not None:
        for sector in projects_data[0]['sectors']:
            print(sector_position, sector)
            new_sheet[sector_position] = sector
            
            sector_items = projects_data[0]['sectors'][sector]
            for item in sector_items:
                # increment cell position to write items under sector
                sector_position = f'{sector_position[0]}{int(sector_position[1:])+1}'
                new_sheet[sector_position] = f'{item["ID"]} {item["name"]}'
                # loop through template again to get positions of quantity, unit and total
                for row in template_sheet.iter_rows():
                    for cell in row:
                        if cell.value == 'Qtd.': new_sheet[f'{cell.coordinate[0]}{sector_position[1:]}'] = item['bom_quant']
                        if cell.value == 'Un': new_sheet[f'{cell.coordinate[0]}{sector_position[1:]}'] = item['unit']
                        if cell.value == 'Total': new_sheet[f'{cell.coordinate[0]}{sector_position[1:]}'] = item['bom_avgprice']
                
                # composition func goes in here
                if item['type'] == 'composicao' and len(item['sector_dict']) != 0: record_composition_data(template_workbook=template_workbook, new_workbook=new_workbook, comp_id=item["ID"], comp_name=item["name"], comp_unit=item['unit'], comp_proj_date=item['Proj_Date'], sector_dict=item['sector_dict'])
                elif item['type'] == 'insumo' and len(item['items_dict']) != 0: last_prod_position_value = record_item_data(template_workbook=template_workbook, new_workbook=new_workbook, prod_id=item['ID'], prod_name=item['name'], prod_unit=item['unit'], items_dict=item['items_dict'], last_prod_position_value=last_prod_position_value)

            # leave a row, then write Total de sectors in next row
            sector_position = f'{sector_position[0]}{int(sector_position[1:])+2}'
            new_sheet[sector_position] = f'Total de {sector}'
            # leave a row again before the next loop
            sector_position = f'{sector_position[0]}{int(sector_position[1:])+2}'


    # take note of last cell position, for use later
    sectors_parameters_position = None
    project_parameters_position = None


    sectors_parameters_list = retrieve_parameters_cell_position_to_list(parameter_value='sectors parameters', template_sheet=template_sheet)
    project_parameters_list = retrieve_parameters_cell_position_to_list(parameter_value='project parameters', template_sheet=template_sheet)

    for project in projects_data:

        for parameter in project_parameters_list:
            project_parameters_position = f'{project_parameters_position[0]}{int(project_parameters_position[1:])+1}'
            new_sheet[project_parameters_position] = f'{parameter}: {project[parameter]}'
        
        project_parameters_position = f'{project_parameters_position[0]}{int(project_parameters_position[1:])+1}'

        for sector in project['sectors']:
            for detail in project['sectors'][sector]:
                for parameter in sectors_parameters_list:
                    sectors_parameters_position = f'{sectors_parameters_position[0]}{int(sectors_parameters_position[1:])+1}'
                    try: new_sheet[sectors_parameters_position] = f'{parameter}: {detail[parameter]}'
                    except: new_sheet[sectors_parameters_position] = ''
                # leave a space for the next set of details in this sector, if any...
                sectors_parameters_position = f'{sectors_parameters_position[0]}{int(sectors_parameters_position[1:])+1}'
            # leave another space for the next sector
            sectors_parameters_position = f'{sectors_parameters_position[0]}{int(sectors_parameters_position[1:])+1}' 


    # save new workbook
    export_report_name = f"report_for_proj_ID_{proj_ID}+item.xlsx"
    new_workbook.save(filename=export_report_name)

    ##################
    #SAVE TO S3 BUCKET
    ##################

    # return "S3/"+export_report_name

if __name__ == '__main__':
    template_xls='template_project_parameters_all_fields.xlsx'
    proj_ID='projects2'
    export_report(template_xls=template_xls, proj_ID=proj_ID)